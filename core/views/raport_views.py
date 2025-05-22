from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import render, redirect
from users.utils import log_actiune
from core.models import Ferma, Fabrica, ProductieFerma, ProductieFabrica, TestCalitateFabrica, TestCalitateFerma
from django.utils.dateparse import parse_date
import csv
import io
import openpyxl
from openpyxl.utils import get_column_letter

def raport_view(request):
    if not (request.user.is_superuser or request.user.face_rapoarte or request.user.rol == 'admin'):
        raise PermissionDenied("Nu ai permisiunea să creezi rapoarte.")

    if request.method == "POST":
        tip = request.POST['tip_raport']
        data_start = request.POST.get('data_start')
        data_end = request.POST.get('data_end')
        format_fisier = request.POST['format']

        # Transforma stringurile in date
        data_start = parse_date(data_start) if data_start else None
        data_end = parse_date(data_end) if data_end else None

        model_map = {
            'ferme': Ferma,
            'fabrici': Fabrica,
            'productie_ferme': ProductieFerma,
            'productie_fabrici': ProductieFabrica,
            'teste_ferme': TestCalitateFerma,
            'teste_fabrici': TestCalitateFabrica,
        }

        model = model_map.get(tip)
        if not model:
            return HttpResponse("Tip raport invalid", status=400)

        queryset = model.objects.all()

        # Filtrare dupa date de adaugare doar dacă modelul are câmpuri de dată, adica tipul nu este Ferma sau Fabrica
        if tip not in ['ferme', 'fabrici']:
            date_fields = [f.name for f in model._meta.fields if
                           'data' in f.name.lower() and f.get_internal_type() in ['DateField', 'DateTimeField']]
            if date_fields and data_start and data_end:
                date_field = date_fields[0]
                queryset = queryset.filter(**{f"{date_field}__range": [data_start, data_end]})

        # Extrage datele sub formă de listă de dict-uri
        rows = list(queryset.values())
        if not rows:
            messages.warning(request, "Nicio înregistrare găsită.")
            return redirect('creare_raport')

        # Extrage header dar nu afiseaza si id
        header = [key for key in rows[0].keys() if key != 'id']

# VERIFICA FORMATUL :

        # FORMAT CSV
        if format_fisier == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{tip}_raport.csv"'

            writer = csv.DictWriter(response, fieldnames=header)
            writer.writeheader()
            writer.writerows(rows)
            return response

        # FORMAT EXCEL (.xlsx)
        elif format_fisier == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Raport"

            # Header
            for col_num, column_title in enumerate(header, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = column_title

            # Continutul randurilor din tabel
            for row_num, row in enumerate(rows, 2):
                for col_num, column_title in enumerate(header, 1):
                    ws.cell(row=row_num, column=col_num, value=row[column_title])

            # Salveaza in memorie fisierul
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{tip}_raport.xlsx"'
            return response

        # FORMAT PDF
        elif format_fisier == 'pdf':
            from django.template.loader import render_to_string
            try:
                from weasyprint import HTML
            except ImportError:
                return HttpResponse("WeasyPrint nu este instalat.", status=500)

            # randeaza html pentru pdf folosid template-ul
            html_string = render_to_string('core/raport_template_pdf.html', {
                'header': header,
                'rows': rows,
            })

            # face conversia spre pdf
            pdf_file = HTML(string=html_string).write_pdf()

            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{tip}_raport.pdf"'
            return response

        else:
            messages.warning(request, "Format invalid!")
            return redirect('creare_raport')

    user = request.user
    log_actiune(request.user, f"{user.username} a creat un raport.")
    return render(request, 'core/creare_raport.html')